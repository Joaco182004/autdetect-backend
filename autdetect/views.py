# Django Imports

from django.http import JsonResponse
from io import BytesIO
from django.conf import settings
from django.contrib.auth.models import User
from django.core.mail import EmailMessage, send_mail
from django.db.models import Case, Count, IntegerField, When
from django.db.models.functions import ExtractMonth, ExtractYear
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone, crypto
from django.utils.crypto import get_random_string
from autdetect.ml_models.ml_model_loader import modelo_tea
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# Django REST Framework Imports
from rest_framework import status, viewsets
from rest_framework.authtoken.models import Token
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.authentication import TokenAuthentication
from rest_framework.response import Response

# Model Imports
from .models import InfantPatient, Psychologists, Questionnaire, UserProfile

# Serializer Imports
from .serializer import InfantPatientSerializer, PsychologistSerializer, QuestionnaireSerializer, UserProfileSerializer, UserSerializer

# ReportLab Imports
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

# Python Standard Library Imports
import os
import random

def enviar_correo(user):
    print(user)
    activation_key = get_random_string(40)
    user_profile = UserProfile.objects.create(user=user, activation_key=activation_key)
    
    activation_url = f"http://localhost:5173/activate/{activation_key}"
    
    subject = "¡Bienvenido/a a AutDetect 🙌 ! Completa tu Registro"
    html_message = f"""
    <html>
    <head></head>
    <body>
    <p>Hola,{user.email}</p>
    <p>Gracias por unirte a <strong>AutDetect</strong>, la plataforma dedicada a la detección temprana de autismo. Nos alegra que formes parte de nuestra comunidad de profesionales comprometidos con el bienestar de los niños y sus familias. 😊</p>
    <p>Para completar tu registro, por favor sigue estos pasos:</p>
    <ol>
        <li><strong>Verifica tu correo electrónico</strong><br>
        Haz clic en el siguiente enlace para confirmar tu dirección de correo electrónico:  <a href="{activation_url}">Verificar correo</a>🔗</li>
        <li><strong>Explora nuestra aplicación web</strong><br>
        Navega por la aplicación web y sus diferentes funcionalidades para comenzar a utilizar la plataforma en tus evaluaciones. 🛠️</li>
    </ol>
    <p>Gracias por unirte a <strong>AutDetect</strong>!<br>
    Estamos emocionados de contar contigo en nuestra misión de hacer la diferencia. 🌟</p>
    <p>Atentamente,<br>
    <strong>AutDetect</strong><br>
    [autdetect@gmail.com] ✉️</p>
    </body>
    </html>
    """
    from_email = settings.DEFAULT_FROM_EMAIL
    recipient_list = [user.email]

    try:
        email = EmailMessage(
            subject=subject,
            body=html_message,
            from_email=from_email,
            to=recipient_list,
        )
        email.content_subtype = 'html'  # Importante para enviar HTML
        email.send()
    except Exception as e:
        print(f"Error al enviar correo: {e}")

@api_view(['GET'])
def activate_account(request, activation_key):
    activation_key = activation_key.rstrip('/')
    try:
        profile = UserProfile.objects.get(activation_key=activation_key)
        if profile.key_expires < timezone.now():
            return HttpResponse('El enlace ha expirado.', status=400)
        user = profile.user
        user.is_active = True
        user.save()
        return HttpResponse('Cuenta activada con éxito. Puedes iniciar sesión.')
    except UserProfile.DoesNotExist:
        return HttpResponse('Enlace inválido.', status=400)

@api_view(['POST'])
def login(request):
    username = request.data.get('username', '')
    password = request.data.get('password', '')

    user = User.objects.filter(username=username).first()
    print(user)
    if not user:
        return Response({"error": "El usuario no está registrado."}, status=status.HTTP_400_BAD_REQUEST)
    if not user.is_active:
        return Response({"error": "El usuario no está activo."}, status=status.HTTP_400_BAD_REQUEST)
    if not user.check_password(password):
        return Response({"error": "El correo o contraseña no son correctos."}, status=status.HTTP_400_BAD_REQUEST)

    token, created = Token.objects.get_or_create(user=user)
    serializer = UserSerializer(instance=user)

    return Response({"token": token.key, "user": serializer.data}, status=status.HTTP_200_OK)

@api_view(['POST'])
def register(request):
    array_errors= []
    print(request.data)
    username = request.data.get('email', '')
    # Buscar el usuario existente
    existing_user_email = User.objects.filter(username=username).first()
    existing_user_dni = Psychologists.objects.filter(dni=request.data.get('dni', '')).first()
    existing_user_tuition_number = Psychologists.objects.filter(tuition_number=request.data.get('tuition_number', '')).first()


    if existing_user_email:
        if existing_user_email.is_active == False:
            Token.objects.filter(user=existing_user_email).delete()
            Psychologists.objects.filter(user=existing_user_email).delete()
            existing_user_email.delete()
        else:
            array_errors.append("El correo electrónico ya se encuentra registrado en el sistema")
    if existing_user_dni:
        array_errors.append("El número de DNI ya se encuentra registrado en el sistema")
    if existing_user_tuition_number:
        array_errors.append("El número de colegiatura ya se encuentra registrado en el sistema")

    if array_errors:
        return Response({"errors": array_errors}, status=status.HTTP_400_BAD_REQUEST)

    # Crear un nuevo usuario
    user_serializer = UserSerializer(data=request.data)
    if user_serializer.is_valid():
        user = user_serializer.save()   
        user.set_password(request.data.get('password', ''))
        user.is_active = False  # Desactivar la cuenta hasta la verificación
        user.save()
    else:
        return Response(user_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # Crear un nuevo token para el usuario
    token, created = Token.objects.get_or_create(user=user)

    # Datos del psicólogo
    psychologist_data = {
        'user': user.id,
        'full_name': request.data.get('full_name', ''),
        'dni': request.data.get('dni', ''),
        'tuition_number': request.data.get('tuition_number', ''),
        'email': request.data.get('email', ''),
    }
    # Crear un nuevo psicólogo
    psychologist_serializer = PsychologistSerializer(data=psychologist_data)

    if psychologist_serializer.is_valid():
        psychologist_serializer.save()
        enviar_correo(user)  # Enviar correo con el enlace de verificación
        return Response({
            'token': token.key,
            'user': user_serializer.data,
            'psychologist': psychologist_serializer.data
        }, status=status.HTTP_201_CREATED)
    else:
        return Response(psychologist_serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class PsychologistView(viewsets.ModelViewSet):
    serializer_class = PsychologistSerializer
    queryset = Psychologists.objects.all()

@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
class InfantPatientView(viewsets.ModelViewSet):
    serializer_class = InfantPatientSerializer
    queryset = InfantPatient.objects.all()
    
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
class QuestionnaireView(viewsets.ModelViewSet):
    serializer_class = QuestionnaireSerializer
    queryset = Questionnaire.objects.all()

@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
class UserProfileView(viewsets.ModelViewSet):
    serializer_class = UserProfileSerializer
    queryset = UserProfile.objects.all()

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def patients_by_month(request):
    year = request.query_params.get('year')
    if not year:
        return Response({"error": "El parámetro 'year' es requerido."}, status=status.HTTP_400_BAD_REQUEST)
    
    monthly_data = (
        Questionnaire.objects
        .filter(date_evaluation__year=year)
        .annotate(
            month=ExtractMonth('date_evaluation')
        )
        .values('month', 'patient__psychology')
        .annotate(total=Count('id'))
        .order_by('month', 'patient__psychology')
    )
    return Response(monthly_data)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def patients_by_month_autism(request):
    year = request.query_params.get('year')
    monthly_data = (
        Questionnaire.objects
        .filter(date_evaluation__year=year)
        .annotate(
            year=ExtractYear('date_evaluation'),  
            month=ExtractMonth('date_evaluation')
        )
        .values('year', 'month','patient__psychology')
        .annotate(
            paciente_con_DT=Count(Case(When(result=0, then=1), output_field=IntegerField())),
            paciente_con_TEA=Count(Case(When(result=1, then=1), output_field=IntegerField())),
        )
        .order_by('year', 'month','patient__psychology')
    )
    return Response(monthly_data)
    
@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def patients_by_gender(request):
    gender_data = (
        InfantPatient.objects
        .values('gender','psychology')
        .annotate(total=Count('id'))
        .order_by('gender','psychology')
    )
    
    return Response(gender_data)

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def get_evaluation_years(request):
    evaluation_years = Questionnaire.objects.dates('date_evaluation', 'year', order='DESC')
    years = [year.year for year in evaluation_years]
    return JsonResponse({'years': years})

def change_email(user,email_change):
    user_profile = UserProfile.objects.filter(user=user).first()

    verification_code = random.randint(100000, 999999)

    if user_profile:
        user_profile.code_change = verification_code
        user_profile.save()
    
    subject = "AutDetect - Cambio de Correo Electrónico 📧"
    html_message = f"""
    <html>
    <head></head>
    <body>
    <p>Hola, {email_change}</p>
    <p>Hemos recibido una solicitud para cambiar la dirección de correo electrónico asociada a tu cuenta en <strong>AutDetect</strong>. Si has solicitado este cambio, por favor confirma tu nueva dirección de correo electrónico utilizando el código de verificación que se muestra a continuación.</p>
    <p><strong>Código de verificación: {verification_code}</strong></p>
    <p>Si no solicitaste este cambio, por favor ignora este correo o contacta a nuestro equipo de soporte.</p>
    <p>Gracias por formar parte de <strong>AutDetect</strong> y por tu dedicación en la detección temprana del autismo. Tu compromiso con esta causa es muy valioso para nosotros. 🌟</p>
    <p>Atentamente,<br>
    <strong>AutDetect</strong><br>
    [autdetect@gmail.com] ✉️</p>
    </body>
    </html>
    """
    from_email = settings.DEFAULT_FROM_EMAIL
    recipient_list = [email_change]

    try:
        email = EmailMessage(
            subject=subject,
            body=html_message,
            from_email=from_email,
            to=recipient_list,
        )
        email.content_subtype = 'html'  # Importante para enviar HTML
        email.send()
    except Exception as e:
        print(f"Error al enviar correo: {e}")

@api_view(['POST'])
def change_email_verification(request):
    username = request.data.get('email', '')
    existing_user = User.objects.filter(username=username).first()
    email_change = request.data.get('email_change', '')
    
    change_email(existing_user,email_change)

    return Response(status=status.HTTP_200_OK)

@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def change_username(request):
    username = request.data.get('email', '')
    existing_user = User.objects.filter(username=username).first()
    email_change = request.data.get('email_change', '')
    if existing_user:
        existing_user.username = email_change
        existing_user.email = email_change
        existing_user.save()
        psychology = Psychologists.objects.filter(email=username).first()
        if psychology:
            psychology.email = email_change
            psychology.save()
    return Response(status=status.HTTP_200_OK)

@api_view(['POST'])
def change_password(request):
    username = request.data.get('email', '')
    existing_user = User.objects.filter(username=username).first()
    password = request.data.get('password', '')
    if existing_user:
        existing_user.set_password(password)
        existing_user.save()
    return Response(status=status.HTTP_200_OK)

@api_view(['POST'])
def change_password_email(request):
    username = request.data.get('email', '')
    
    existing_user = User.objects.filter(username=username).first()
    
    user_profile = UserProfile.objects.filter(user=existing_user).first()

    verification_code = random.randint(100000, 999999)

    if user_profile:
        user_profile.code_change = verification_code
        user_profile.save()
    else:
        return Response("El correo ingresado no se encuentra registrado.", status=status.HTTP_400_BAD_REQUEST)
    
    subject = "AutDetect - Cambio de Contraseña 🔑"
    html_message = f"""
    <html>
    <head></head>
    <body>
    <p>Hola, {username}</p>
    <p>Hemos recibido una solicitud para cambiar la contraseña asociada a tu cuenta en <strong>AutDetect</strong>. Si has solicitado este cambio, por favor confirma tu nueva contraseña utilizando el código de verificación que se muestra a continuación.</p>
    <p><strong>Código de verificación: {verification_code}</strong></p>
    <p>Si no solicitaste este cambio, por favor ignora este correo o contacta a nuestro equipo de soporte.</p>
    <p>Gracias por formar parte de <strong>AutDetect</strong> y por tu dedicación en la detección temprana del autismo. Tu compromiso con esta causa es muy valioso para nosotros. 🌟</p>
    <p>Atentamente,<br>
    <strong>AutDetect</strong><br>
    [autdetect@gmail.com] ✉️</p>
    </body>
    </html>
    """
    from_email = settings.DEFAULT_FROM_EMAIL
    recipient_list = [username]

    try:
        email = EmailMessage(
            subject=subject,
            body=html_message,
            from_email=from_email,
            to=recipient_list,
        )
        email.content_subtype = 'html'  # Importante para enviar HTML
        email.send()
        return Response(status=status.HTTP_200_OK)
    except Exception as e:
        print(f"Error al enviar correo: {e}")


@api_view(['POST'])
def validate_code(request):
    username = request.data.get('email', '')
    code = request.data.get('code', '')
    
    existing_user = User.objects.filter(username=username).first()

    if existing_user:
        user_profile = UserProfile.objects.filter(user=existing_user).first()
        if(user_profile.code_change == code):
            return Response({"success": True}, status=status.HTTP_200_OK)
        else:
            return Response("El código no corresponde al código enviado.", status=status.HTTP_400_BAD_REQUEST)

def transform_yes_no(value):
    return "Sí" if value == 1 else "No"

def generar_reporte_pdf(test):
    patient = test.patient
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte.pdf"'
    buffer = BytesIO()
    # Crear el documento PDF
    doc = SimpleDocTemplate(buffer, pagesize=A4)

    # Obtener el estilo para los párrafos
    styles = getSampleStyleSheet()

    # Contenido del PDF (secciones)
    elements = []

    # Ruta del logo (asegúrate de colocar la ruta correcta)
    logo_path = os.path.join('crudautdetect/static/imgs/logoautdetect.png')

    logo = Image(logo_path, 100, 100)
    logo.hAlign = 'CENTER'  # Alineación centrada del logo
    
    # Crear los encabezados de texto y centrar
    header1 = Paragraph("AutDetect", styles['Normal'])
    header2 = Paragraph("Contacto: autdetect@gmail.com", styles['Normal'])
    
    header1.hAlign = 'CENTER'  # Centrar texto
    header2.hAlign = 'CENTER'  # Centrar texto
    
    # Añadir el logo centrado
    elements.append(logo)
    
    # Añadir un espacio
    
    elements.append(Paragraph("Reporte de detección temprana del Trastorno Espectro Autista", styles['Title']))

    # Subtítulo
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("1. Respuestas del cuestionario de comportamiento", styles['Heading2']))

    # Preguntas
    questions = [
        "1. ¿Responde tu hijo(a) cuando lo(a) llamas por su nombre, como volteándose, hablándote o dejando de hacer lo que estaba haciendo?",
        "2. ¿Te mira tu hijo(a) a los ojos cuando le hablas, juegas con él/ella, o lo(a) vistes?",
        "3. ¿Tu hijo(a) señala con el dedo o la mano cuando quiere algo o necesita ayuda, como un juguete o comida que no puede alcanzar?",
        "4. ¿Alguna vez tu hijo(a) señala algo que le causa interés solo para mostrarte, como un avión en el cielo o un animal?",
        "5. ¿Tu hijo(a) juega a hacer cosas como beber de una taza de juguete, hablar por teléfono, o darle de comer a una muñeca o peluche?",
        "6. Si te giras a ver algo, ¿tu hijo(a) trata de mirar hacia lo que estás mirando?",
        "7. Si tú o alguien más en la familia está visiblemente triste o molesto, ¿tu hijo(a) muestra signos de querer consolarlo?",
        "8. ¿Tu hijo(a) dijo sus primeras palabras (como 'mamá' o 'papá') alrededor del primer año de vida?",
        "9. ¿Usa tu hijo(a) gestos como decir adiós con la mano, aplaudir, o imitar algún sonido gracioso que haces?",
        "10. ¿Ha notado que su hijo(a) se queda mirando un objeto o al vacío durante un tiempo prolongado, sin parecer darse cuenta de lo que ocurre a su alrededor?",
        "11. ¿Tu hijo(a) ha presentado alguna vez ictericia, es decir, un tono amarillento en la piel o en los ojos, especialmente poco después de nacer?",
        "12. ¿Hay algún familiar en tu familia que haya sido diagnosticado con Trastorno del Espectro Autista (TEA)?"
    ]

    answers = [
        transform_yes_no(1 if test.pregunta_1 == 0 else 0),
        transform_yes_no(1 if test.pregunta_2 == 0 else 0),
        transform_yes_no(1 if test.pregunta_3 == 0 else 0),
        transform_yes_no(1 if test.pregunta_4 == 0 else 0),
        transform_yes_no(1 if test.pregunta_5 == 0 else 0),
        transform_yes_no(1 if test.pregunta_6 == 0 else 0),
        transform_yes_no(1 if test.pregunta_7 == 0 else 0),
        transform_yes_no(1 if test.pregunta_8 == 0 else 0),
        transform_yes_no(1 if test.pregunta_9 == 0 else 0),
        transform_yes_no(test.pregunta_10),  # Cociente Espectro Autista
        transform_yes_no(test.ictericia),
        transform_yes_no(test.familiar_con_tea)
    ]
    
    for question, answer in zip(questions, answers):
        elements.append(Spacer(1, 5))  # Espacio entre elementos
        elements.append(Paragraph(question, styles['Normal']))  # Agregar la pregunta
        elements.append(Spacer(1, 2)) 
        elements.append(Paragraph(f"<b>Respuesta:</b> {answer}", styles['Normal']))

    # Agregar un salto de página
    elements.append(Spacer(1, 12))

    # Subtítulo de la segunda sección
    elements.append(Paragraph("2. Resultados de la evaluación", styles['Heading2']))
    
    elements.append(Spacer(1, 12))
    # Datos de la tabla
    table_data = [
        ["DNI","Nombre", "Fecha de nacimiento", "Fecha de evaluación", "Resultado", "Probabilidad"],
        [patient.infant_dni,patient.infant_name, patient.birth_date, test.date_evaluation, "Positivo" if test.result == 1 else "Negativo", str((int(test.probability * 100) / 100.0) * 100) + "%"],
    ]

    # Crear la tabla
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.skyblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    # Añadir la tabla a los elementos
    elements.append(table)

    # Construir el documento PDF con el contenido
    doc.build(elements)

    buffer.seek(0)
    return buffer

@api_view(['POST'])
def send_email_report(request):
    id_test = request.data.get('id_test', '')
    try:
        questionnaire = Questionnaire.objects.get(id=id_test)
        patient = questionnaire.patient
        pdf_buffer = generar_reporte_pdf(questionnaire)
        subject = "AutDetect - Reporte de Evaluación 📄"
        html_message = f"""
        <html>
        <head></head>
        <body>
        <p>Hola, {patient.guardian_name}</p>
        <p>Adjunto encontrarás el reporte de evaluación de AutDetect en formato PDF de su hijo(a) {patient.infant_name}.</p>
        <p>Si tienes alguna pregunta o necesitas asistencia adicional, no dudes en ponerte en contacto con nuestro equipo.</p>
        <p>Gracias por tu colaboración en la detección temprana del autismo.</p>
        <p>Atentamente,<br>
        <strong>AutDetect</strong><br>
        [autdetect@gmail.com] ✉️</p>
        </body>
        </html>
        """
        from_email = settings.DEFAULT_FROM_EMAIL
        recipient_list = [patient.guardian_email]

        try:
            email = EmailMessage(
                subject=subject,
                body=html_message,
                from_email=from_email,
                to=recipient_list,
            )
            email.attach('reporte_autdetect_'+patient.infant_name+'.pdf', pdf_buffer.getvalue(), 'application/pdf')

            email.content_subtype = 'html'  # Importante para enviar HTML
            email.send()
            return Response(status=status.HTTP_200_OK)
        except Exception as e:
            print(f"Error al enviar correo: {e}")
    except Questionnaire.DoesNotExist:
        # Manejo si no existe el cuestionario
        return None
    

# Modelo ML
@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def prediccion_view(request):
    try:
        # Obtener datos de entrada desde el cuerpo de la solicitud JSON
        data = request.data

        # Extraer las características del objeto JSON
        pregunta_1 = int(data.get('pregunta_1'))
        pregunta_2 = int(data.get('pregunta_2'))
        pregunta_3 = int(data.get('pregunta_3'))
        pregunta_4 = int(data.get('pregunta_4'))
        pregunta_5 = int(data.get('pregunta_5'))
        pregunta_6 = int(data.get('pregunta_6'))
        pregunta_7 = int(data.get('pregunta_7'))
        pregunta_8 = int(data.get('pregunta_8'))
        pregunta_9 = int(data.get('pregunta_9'))
        pregunta_10 = int(data.get('pregunta_10_Cociente_Espectro_Autista'))
        sexo = int(data.get('Sexo'))
        ictericia = int(data.get('Ictericia'))
        familiar_con_tea = int(data.get('Familiar_con_TEA'))

        # Crear el arreglo de entrada para el modelo
        datos_de_entrada = np.array([[pregunta_1, pregunta_2, pregunta_3, pregunta_4, pregunta_5,
                                      pregunta_6, pregunta_7, pregunta_8, pregunta_9, pregunta_10,
                                      sexo, ictericia, familiar_con_tea]])

        # Hacer predicción
        prediccion = modelo_tea.predict(datos_de_entrada)[0]
        probabilidad = modelo_tea.predict_proba(datos_de_entrada)[0][1]  # Probabilidad de la clase positiva

        # Devolver la respuesta como JSON
        return Response({'prediccion': int(prediccion), 'probabilidad': float(probabilidad)}, status=status.HTTP_200_OK)

    except Exception as e:
        # Manejar posibles errores
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
#Reportes
def ajustar_ancho_columna(ws):
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter  # Obtener la letra de la columna
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # Ajustar un poco más el ancho
        ws.column_dimensions[column].width = adjusted_width

def aplicar_estilo_encabezado(ws):
    # Definir los estilos
    font = Font(bold=True, color="FFFFFF")  # Negrita y texto blanco
    fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Fondo azul cielo
    center_alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:  # La primera fila es el encabezado
        cell.font = font
        cell.fill = fill
        cell.alignment = center_alignment

def aplicar_bordes(ws):
    # Definir los bordes
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Aplicar los bordes a todas las celdas con texto
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:  # Aplicar borde solo si la celda tiene valor
                cell.border = thin_border
                cell.alignment=center_alignment

# Vista para exportar los datos de InfantPatient
@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def export_infant_patients_excel(request,id_psychology):
    # Crear el archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Infant Patients'

    # Escribir los encabezados
    ws.append(['DNI del Paciente', 'Nombre del Paciente', 'Fecha de Nacimiento', 'Género', 'DNI del Tutor', 'Nombre del Tutor', 'Correo del Tutor', 'Teléfono de Contacto', 'Distrito', 'Psicólogo'])

    aplicar_estilo_encabezado(ws)
    # Obtener los datos
    patients = InfantPatient.objects.all()

    # Escribir los datos
    for patient in patients:
        if patient.psychology.id == id_psychology:
            ws.append([
                patient.infant_dni,
                patient.infant_name,
                patient.birth_date,
                patient.get_gender_display(),
                patient.guardian_dni,
                patient.guardian_name,
                patient.guardian_email,
                patient.contact_phone,
                patient.district,
                patient.psychology.full_name,  # Asumiendo que 'Psychologists' tiene un campo 'name'
            ])

    ajustar_ancho_columna(ws)
    aplicar_bordes(ws)
    # Preparar la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=infant_patients.xlsx'
    
    # Guardar el archivo en la respuesta
    wb.save(response)
    return response


# Vista para exportar los datos de Questionnaire
@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def export_questionnaires_excel(request,id_psychology):
    # Crear el archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Questionnaires'

    # Escribir los encabezados
    ws.append(['DNI del paciente','Nombre del Paciente', 'Psicólogo', 'Pregunta 1', 'Pregunta 2', 'Pregunta 3', 'Pregunta 4', 'Pregunta 5', 'Pregunta 6', 'Pregunta 7', 'Pregunta 8', 'Pregunta 9', 'Pregunta 10 Cociente Espectro Autista', 'Ictericia', 'Familiar con TEA', 'Resultado', 'Probabilidad', 'Fecha de Evaluación'])

    aplicar_estilo_encabezado(ws)
    # Obtener los datos
    questionnaires = Questionnaire.objects.all()

    # Escribir los datos
    for q in questionnaires:
        if q.patient.psychology.id == id_psychology:
            ws.append([
            q.patient.infant_dni,
            q.patient.infant_name,
            q.patient.psychology.full_name,
            str(q.pregunta_1),  # Asegúrate de que se traten como texto
            str(q.pregunta_2),
            str(q.pregunta_3),
            str(q.pregunta_4),
            str(q.pregunta_5),
            str(q.pregunta_6),
            str(q.pregunta_7),
            str(q.pregunta_8),
            str(q.pregunta_9),
            str(q.pregunta_10),
            str(q.ictericia),
            str(q.familiar_con_tea),
            'Sí' if q.result else 'No',
            str((int(q.probability * 100) / 100.0) * 100) + "%",
            q.date_evaluation,
            ])

    ajustar_ancho_columna(ws)
    aplicar_bordes(ws)
    # Preparar la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=questionnaires.xlsx'
    
    # Guardar el archivo en la respuesta
    wb.save(response)
    return response

# Exportar en un pdf el reporte

@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def export_report_evaluation(request):
    id_test = request.data.get('id_test', '')
    try:
        questionnaire = Questionnaire.objects.get(id=id_test)
        patient = questionnaire.patient
        pdf_buffer = generar_reporte_pdf(questionnaire)
        response = HttpResponse(pdf_buffer.getvalue(),content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_autdetect_'+patient.infant_name+'.pdf"'
        return response
    except Questionnaire.DoesNotExist:
        # Manejo si no existe el cuestionario
        return None
